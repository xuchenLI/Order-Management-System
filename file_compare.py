# file_compare.py

import os
import pandas as pd
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def open_file_compare_tool():
    """
    比较两个 Excel 文件：
      - 以 A 的行顺序为基准，把与 B 相同列 (除sku cls外) 以并排列显示；
      - 不一致的单元格标黄色；
      - A有但B没有的整行标蓝色；
      - B有但A没有的行放到第二个 Sheet 中列出来。
    """
    try:
        options = QFileDialog.Option.DontUseNativeDialog

        # 1) 让用户选择 Excel A
        file_a, _ = QFileDialog.getOpenFileName(
            None,
            "选择 Excel A (基准)",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options
        )
        if not file_a:
            return

        # 2) 让用户选择 Excel B
        file_b, _ = QFileDialog.getOpenFileName(
            None,
            "选择 Excel B (待比对)",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options
        )
        if not file_b:
            return

        # 3) 用 pandas 分别读取 A / B
        dfA = pd.read_excel(file_a)
        dfB = pd.read_excel(file_b)

        # 将列名统一转小写、去除多余空格，便于后续匹配
        dfA.columns = dfA.columns.str.strip().str.lower()
        dfB.columns = dfB.columns.str.strip().str.lower()

        # 4) 确保 A 和 B 都含 'sku cls' 列
        if 'sku cls' not in dfA.columns:
            QMessageBox.warning(None, "错误", "Excel A 中缺少 'SKU CLS' 列。")
            return
        if 'sku cls' not in dfB.columns:
            QMessageBox.warning(None, "错误", "Excel B 中缺少 'SKU CLS' 列。")
            return

        # 统一 SKU CLS 格式
        dfA['sku cls'] = dfA['sku cls'].astype(str).str.lower().str.strip()
        dfB['sku cls'] = dfB['sku cls'].astype(str).str.lower().str.strip()

        # 5) 找出 B 中有，但 A 中没有的行 (ExtraInB)
        #    以后会放在第二个 Sheet
        sku_in_A = set(dfA['sku cls'].unique())
        sku_in_B = set(dfB['sku cls'].unique())
        extra_in_B_skus = sku_in_B - sku_in_A
        dfB_extra = dfB[dfB['sku cls'].isin(extra_in_B_skus)].copy()

        # 6) 构建一个以 B 的 sku cls 为键的 dict，便于快速查行
        dfB_dict = {}
        for idx, row in dfB.iterrows():
            sku_b = row['sku cls']
            dfB_dict[sku_b] = row

        # 要对比 A、B 共同拥有的字段（除 'sku cls' 外）
        common_cols = [
            col for col in dfA.columns
            if col in dfB.columns and col != 'sku cls'
        ]

        # 7) 在 A 的 DataFrame 中插入 B 对应列（形如 col+"_b"）
        dfA_compare = dfA.copy()
        for col in common_cols:
            dfA_compare[col + '_b'] = ""

        # 新增一列记录是否在 B 中找到
        dfA_compare['found_in_B'] = ""

        # 8) 逐行遍历 A，以 A 的顺序写结果
        for idx, rowA in dfA_compare.iterrows():
            sku_a = rowA['sku cls']
            if sku_a in dfB_dict:
                dfA_compare.at[idx, 'found_in_B'] = "YES"
                rowB = dfB_dict[sku_a]
                # 为每个 common col 填写B中的值
                for col in common_cols:
                    dfA_compare.at[idx, col + '_b'] = str(rowB[col])
            else:
                dfA_compare.at[idx, 'found_in_B'] = "NO"

        # 9) 重新排一下列顺序，使得:
        #    [sku cls, (col, col_b, col2, col2_b, ...), 其余列, found_in_B]
        original_cols = list(dfA.columns)  # [sku cls, col1, col2, ...]
        final_cols = []

        # 先加 'sku cls'
        if 'sku cls' in original_cols:
            final_cols.append('sku cls')

        for c in original_cols:
            if c == 'sku cls':
                continue
            if c in common_cols:
                final_cols.append(c)         # A 的列
                final_cols.append(c + '_b')  # B 的列
            else:
                final_cols.append(c)

        # 最后加上 found_in_B
        final_cols.append('found_in_B')

        # 过滤一下，确保列存在
        final_cols = [col for col in final_cols if col in dfA_compare.columns]
        dfA_compare = dfA_compare[final_cols]

        # 10) 将 dfA_compare 存到临时文件
        temp_file = 'temp_compare.xlsx'
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            dfA_compare.to_excel(writer, sheet_name='CompareResult', index=False)
            # 同时把 dfB_extra 写到第二个 sheet
            if not dfB_extra.empty:
                dfB_extra.to_excel(writer, sheet_name='ExtraInB', index=False)

        # 11) 用 openpyxl 打开该临时文件，对 CompareResult 做高亮
        wb = load_workbook(temp_file)
        ws_compare = wb['CompareResult']

        # 如果存在 extraInB 行，则不需要特别高亮，因为它是在第二个 sheet
        # 你也可以对那部分再做其他颜色标记

        # 颜色填充
        highlight_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        highlight_blue = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')  # 浅蓝

        # 构建列名 -> 列号映射
        all_cols = list(dfA_compare.columns)
        col_map = {col_name: i+1 for i, col_name in enumerate(all_cols)}  # +1 因为 Excel 列从 1 开始

        # 找出可对比的 (col, col+"_b") pairs
        compare_pairs = []
        for col in common_cols:
            if col in dfA_compare.columns and (col + '_b') in dfA_compare.columns:
                compare_pairs.append((col, col + '_b'))

        # 行数
        row_count = len(dfA_compare)
        for i in range(row_count):
            excel_row = i + 2  # 第1行为表头
            found_in_b_value = str(dfA_compare.at[i, 'found_in_B']).strip()

            if found_in_b_value == "NO":
                # A有B没有 => 整行标蓝色
                for col_name in all_cols:
                    col_idx = col_map[col_name]
                    cell = ws_compare.cell(row=excel_row, column=col_idx)
                    cell.fill = highlight_blue
            else:
                # A和B都存在 => 对比不一致的单元格标黄
                for (colA, colB) in compare_pairs:
                    valA = str(dfA_compare.at[i, colA]).strip()
                    valB = str(dfA_compare.at[i, colB]).strip()
                    if valA and valB and (valA != valB):
                        colA_idx = col_map[colA]
                        colB_idx = col_map[colB]
                        cellA = ws_compare.cell(row=excel_row, column=colA_idx)
                        cellB = ws_compare.cell(row=excel_row, column=colB_idx)
                        cellA.fill = highlight_yellow
                        cellB.fill = highlight_yellow

        # 12) 让用户选择最终保存结果的位置
        save_file, _ = QFileDialog.getSaveFileName(
            None,
            "保存对比结果",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options
        )
        if not save_file:
            wb.close()
            os.remove(temp_file)
            return

        if not save_file.endswith('.xlsx'):
            save_file += '.xlsx'

        # 保存结果
        wb.save(save_file)
        wb.close()
        os.remove(temp_file)  # 删除临时文件

        QMessageBox.information(None, "完成", f"对比完成，结果已保存到：\n{save_file}")

    except Exception as e:
        print(f"比较文件时发生错误：{e}")
        QMessageBox.critical(None, "错误", f"比较文件时发生错误：\n{e}")
