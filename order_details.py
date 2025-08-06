# order_details.py

import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QLineEdit, QMessageBox, QGridLayout, QComboBox, QFileDialog, QListWidget, QAbstractItemView, QListWidgetItem, QApplication
)
from PyQt6.QtGui import QShortcut, QKeySequence
from PyQt6.QtCore import Qt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from data import (
    purchase_orders, deleted_orders, db_fields, delete_purchase_order_from_db,
    save_purchase_order_to_db, data_manager, update_inventory, inventory,
    update_inventory_arrival_date, get_purchase_order_by_nb, products, get_product_by_sku, save_product_to_db,
    load_products_from_db, delete_product_from_db, cascade_update_purchase_order_by_guid, get_po_guid_for_inventory,
    sales_orders
)
from price_calculator import open_price_calculator
import json
import datetime, re
import uuid
import os
from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QListWidget, QLineEdit, QFileDialog, QMessageBox

class SupplierManagerDialog(QDialog):
    def __init__(self, parent, supplier_combo, suppliers_dir=r"D:\00_Programming\98_Pycharm\00_Workplace\Order Manager\Official_Tool\01_Cursor_Code\01_Working\00_Main Branch\suppliers"):
        super().__init__(parent)
        self.setWindowTitle("供应商管理")
        self.setGeometry(400, 300, 500, 400)
        self.suppliers_dir = suppliers_dir
        self.supplier_combo = supplier_combo
        if not os.path.exists(self.suppliers_dir):
            os.makedirs(self.suppliers_dir)
        # 自动为下拉框已有供应商补全文件夹
        if hasattr(self.supplier_combo, 'count') and hasattr(self.supplier_combo, 'itemText'):
            for i in range(self.supplier_combo.count()):
                name = self.supplier_combo.itemText(i)
                if name and not os.path.exists(os.path.join(self.suppliers_dir, name)):
                    os.makedirs(os.path.join(self.suppliers_dir, name))
        layout = QVBoxLayout()
        self.setLayout(layout)
        # 供应商列表
        self.list_widget = QListWidget()
        self.refresh_supplier_list()
        layout.addWidget(QLabel("现有供应商："))
        layout.addWidget(self.list_widget)
        # 新增供应商区域
        add_layout = QHBoxLayout()
        self.input_new = QLineEdit(); self.input_new.setPlaceholderText("新供应商名称")
        btn_upload = QPushButton("上传文件")
        btn_add = QPushButton("新添加供应商")
        add_layout.addWidget(self.input_new)
        add_layout.addWidget(btn_upload)
        add_layout.addWidget(btn_add)
        layout.addLayout(add_layout)
        btn_upload.clicked.connect(self.upload_file)
        btn_add.clicked.connect(self.add_supplier)

    def is_valid_supplier_name(self, name):
        # 禁止非法文件名字符
        return bool(name) and not re.search(r'[\\/:*?"<>|]', name)

    def refresh_supplier_list(self):
        self.list_widget.clear()
        suppliers = [name for name in os.listdir(self.suppliers_dir) if os.path.isdir(os.path.join(self.suppliers_dir, name))]
        for name in suppliers:
            item_widget = QWidget()
            hbox = QHBoxLayout()
            hbox.setContentsMargins(0,0,0,0)
            hbox.setSpacing(4)
            label = QLabel(name)
            btn_folder = QPushButton()
            btn_folder.setText("📁")
            btn_folder.setFixedWidth(28)
            btn_folder.clicked.connect(lambda _, n=name: self.open_folder(n))
            hbox.addWidget(label)
            hbox.addWidget(btn_folder)
            hbox.addStretch()
            item_widget.setLayout(hbox)
            from PyQt6.QtWidgets import QListWidgetItem
            item = QListWidgetItem(self.list_widget)
            self.list_widget.addItem(item)
            self.list_widget.setItemWidget(item, item_widget)

    def open_folder(self, supplier_name):
        folder_path = os.path.abspath(os.path.join(self.suppliers_dir, supplier_name))
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        import subprocess
        if os.name == 'nt':
            os.startfile(folder_path)
        elif os.name == 'posix':
            subprocess.Popen(['xdg-open', folder_path])
        else:
            QMessageBox.information(self, "提示", f"请手动打开: {folder_path}")

    def upload_file(self):
        supplier_name = self.input_new.text().strip()
        if not self.is_valid_supplier_name(supplier_name):
            QMessageBox.warning(self, "提示", "请先输入合法的新供应商名称")
            return
        folder_path = os.path.join(self.suppliers_dir, supplier_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        files, _ = QFileDialog.getOpenFileNames(self, "选择要上传的文件")
        for f in files:
            try:
                import shutil
                shutil.copy(f, folder_path)
            except Exception as e:
                QMessageBox.warning(self, "上传失败", f"文件 {f} 上传失败: {e}")
        QMessageBox.information(self, "上传成功", "文件已上传到供应商文件夹")

    def add_supplier(self):
        supplier_name = self.input_new.text().strip()
        if not self.is_valid_supplier_name(supplier_name):
            QMessageBox.warning(self, "提示", "请输入合法的供应商名称（不能包含\\/:*?\"<>|等字符）")
            return
        folder_path = os.path.join(self.suppliers_dir, supplier_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        # 加入下拉框
        if hasattr(self.supplier_combo, 'count') and hasattr(self.supplier_combo, 'itemText'):
            if supplier_name not in [self.supplier_combo.itemText(i) for i in range(self.supplier_combo.count())]:
                self.supplier_combo.addItem(supplier_name)
            self.supplier_combo.setCurrentText(supplier_name)
        self.refresh_supplier_list()
        QMessageBox.information(self, "成功", f"已添加供应商：{supplier_name}")

class OrderDetailsWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("采购订单管理")
        self.setGeometry(200, 200, 1400, 700)
        self.editing_in_progress = False
        self.layout_main = QVBoxLayout()

        # 输入区域
        self.layout_inputs = QGridLayout()
        self.layout_inputs.setHorizontalSpacing(10)
        self.entries = {}

        # 左侧字段
        left_fields = [
            ('订单号', 'Order Nb'),
            ('产品编号', 'Product_ID'),
            ('Order Type', 'Order Type'),
            # ('Order Step', 'Order Step'),  # 删除这一项
            ('期望利润', "Expected Profit"),
            ('境内运费', 'Domestic Freight CAD'),
            ('国际运费', 'International Freight EURO'),
            ('EXW 汇率', 'EXW Exchange Rate'),
            ('国际运费汇率', 'International Freight Exchange Rate'),
            # 新增字段
            ("UCC14", "UCC14"),
            ("UCC13", "UCC13"),
        ]

        
        right_fields = [
            ("Supplier", "Supplier"),
            ("BCMB", "BCMB"),
            ("SKU CLS", "SKU CLS"),
            ("Supplier Order Number", "Supplier Order Number"),
            ("ITEM Name", "ITEM Name"),
            ("CATEGORY", "CATEGORY"),
            ("SIZE", "SIZE"),
            ("ALC.", "ALC."),
            ("QUANTITY CS", "QUANTITY CS"),
            ("BTL PER CS", "BTL PER CS"),
            ("EXW(€)", "EXW EURO"),
            ("REMARKS", "REMARKS"),
        ]

        # 左侧字段
        for row, (label_text, field_name) in enumerate(left_fields):
            label = QLabel(label_text + ":")
            label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if field_name == "Order Type":
                entry = QComboBox()
                entry.addItems(["Allocation", "In Stock"])
                entry.setCurrentText("Allocation")
            elif field_name == "Expected Profit":
                entry = QLineEdit()
                entry.setText("0.05")
            elif field_name == "Domestic Freight CAD":
                entry = QLineEdit()
                entry.setText("35")
            elif field_name == "International Freight EURO":
                entry = QLineEdit()
                entry.setText("0")
            elif field_name == "EXW Exchange Rate":
                entry = QLineEdit()
                entry.setText("0")
            elif field_name == "International Freight Exchange Rate":
                entry = QLineEdit()
                entry.setText("0")
            else:
                entry = QLineEdit()
            entry.setFixedWidth(500)
            entry.setStyleSheet("border: 1px solid lightgray;")
            self.entries[field_name] = entry
            self.layout_inputs.addWidget(label, row, 0)
            self.layout_inputs.addWidget(entry, row, 1)

        for row, (label_text, field_name) in enumerate(right_fields):
            label = QLabel(label_text + ":")
            label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if field_name == "Supplier":
                entry = QComboBox()
                # 动态从suppliers文件夹读取供应商名称，而不是硬编码
                suppliers_dir = r"D:\00_Programming\98_Pycharm\00_Workplace\Order Manager\Official_Tool\01_Cursor_Code\01_Working\00_Main Branch\suppliers"
                if os.path.exists(suppliers_dir):
                    suppliers = [name for name in os.listdir(suppliers_dir) if os.path.isdir(os.path.join(suppliers_dir, name))]
                    entry.addItems(suppliers)
                else:
                    # 如果文件夹不存在，使用默认列表
                    entry.addItems(["Filips", "CVBG", "DULONG", "BONCHAEAU"])
                btn_add_supplier = QPushButton("+")
                btn_add_supplier.setFixedWidth(24)
                def open_supplier_manager():
                    dlg = SupplierManagerDialog(self, entry)
                    dlg.exec()
                btn_add_supplier.clicked.connect(open_supplier_manager)
                supplier_layout = QHBoxLayout()
                supplier_layout.setContentsMargins(0,0,0,0)
                supplier_layout.setSpacing(2)
                supplier_layout.addWidget(entry)
                supplier_layout.addWidget(btn_add_supplier)
                supplier_widget = QWidget()
                supplier_widget.setLayout(supplier_layout)
                self.entries[field_name] = entry
                self.layout_inputs.addWidget(label, row, 2)
                self.layout_inputs.addWidget(supplier_widget, row, 3)
                continue
            elif field_name == "CATEGORY":
                entry = QComboBox()
                entry.addItems(["RED", "WHITE", "ROSE", "CHAMPAGNE"])
                entry.setCurrentText("RED")
            elif field_name == "SIZE":
                entry = QComboBox()
                entry.addItems(["0.75", "1.5", "3", "6", "9"])
                entry.setCurrentText("0.75")
            else:
                entry = QLineEdit()
            entry.setFixedWidth(500)
            entry.setStyleSheet("border: 1px solid lightgray;")
            self.entries[field_name] = entry
            self.layout_inputs.addWidget(label, row, 2)
            self.layout_inputs.addWidget(entry, row, 3)

        self.layout_main.addLayout(self.layout_inputs)

        # 按钮区域
        layout_buttons = QHBoxLayout()

        self.button_add = QPushButton("添加采购订单")
        self.button_add.clicked.connect(self.add_order)

        self.button_update = QPushButton("更新采购订单")
        self.button_update.clicked.connect(self.update_order)

        self.entry_search = QLineEdit()
        self.entry_search.setMaxLength(15)
        self.entry_search.setFixedWidth(100)
        button_search = QPushButton("查找订单")
        button_search.clicked.connect(self.find_order)

        self.entry_delete = QLineEdit()
        self.entry_delete.setMaxLength(30)
        self.entry_delete.setFixedWidth(100)
        button_delete = QPushButton("删除订单")
        button_delete.clicked.connect(self.delete_order)

        undo_button = QPushButton("撤销删除")
        undo_button.clicked.connect(self.undo_delete_order)

        button_export = QPushButton("导出订单")
        button_export.clicked.connect(self.export_orders)

        button_price_calculator = QPushButton("价格计算器")
        button_price_calculator.clicked.connect(lambda: self.open_price_calculator())

        button_compare = QPushButton("比较注册信息")
        button_compare.clicked.connect(self.compare_with_registration_file)

        layout_buttons.addWidget(self.button_add)
        layout_buttons.addWidget(self.button_update)
        layout_buttons.addWidget(button_search)
        layout_buttons.addWidget(self.entry_search)
        layout_buttons.addWidget(button_delete)
        layout_buttons.addWidget(self.entry_delete)
        layout_buttons.addWidget(undo_button)
        layout_buttons.addWidget(button_export)
        layout_buttons.addWidget(button_price_calculator)
        layout_buttons.addWidget(button_compare)

        self.layout_main.addLayout(layout_buttons)

        # 增加排序和过滤区域
        filter_sort_layout = QHBoxLayout()

        sort_label = QLabel("排序规则:")
        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["按更新时间", "按订单号", "按Product_ID", "按SKU CLS", "按Supplier", "按ITEM Name", "按CATEGORY", "按INVOICE Price", "按WHOLESALE CS", "按EXW"])
        self.sort_combo.currentIndexChanged.connect(self.update_order_table)
        filter_sort_layout.addWidget(sort_label)
        filter_sort_layout.addWidget(self.sort_combo)

        filter_field_label = QLabel("过滤条件:")
        self.filter_field_combo = QComboBox()
        self.filter_field_combo.addItems(["按订单号", "按Supplier", "按ITEM Name", "按CATEGORY", "按Date", "按SKU CLS"])
        filter_sort_layout.addWidget(filter_field_label)
        filter_sort_layout.addWidget(self.filter_field_combo)

        self.filter_field_input = QLineEdit()
        self.filter_field_input.setPlaceholderText("输入过滤内容(支持*通配符)")
        self.filter_field_input.textChanged.connect(self.update_order_table)
        filter_sort_layout.addWidget(self.filter_field_input)

        self.layout_main.addLayout(filter_sort_layout)

        # 订单列表显示区域
        display_fields = [field for field in db_fields]
        self.order_table = QTableWidget()
        self.order_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        self.order_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        # 将编辑触发模式设为双击或 F2 编辑，而非完全只读
        self.order_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.EditKeyPressed)
        self.order_table.setColumnCount(len(display_fields))
        self.order_table.setHorizontalHeaderLabels([label_text for label_text, field_name in display_fields])
        self.order_table.verticalHeader().setVisible(False)
        self.order_table.horizontalHeader().setStretchLastSection(True)
        self.order_table.setWordWrap(True)
        self.order_table.resizeColumnsToContents()
        # 连接订单表格的选择信号
        self.order_table.selectionModel().selectionChanged.connect(self.on_order_selected)
        # 连接编辑完成信号
        self.order_table.itemChanged.connect(self.on_item_changed)
        self.layout_main.addWidget(self.order_table)
        # 添加复制快捷键（Ctrl+C）和 ESC 快捷键取消选择
        self.copy_shortcut = QShortcut(QKeySequence(QKeySequence.StandardKey.Copy), self.order_table)
        self.copy_shortcut.activated.connect(self.copySelectedCells)
        self.clear_selection_shortcut = QShortcut(QKeySequence("Escape"), self.order_table)
        self.clear_selection_shortcut.activated.connect(lambda: self.order_table.clearSelection())
        self.setLayout(self.layout_main)
        self.update_order_table()

    def copySelectedCells(self):
        selected_ranges = self.order_table.selectedRanges()
        if not selected_ranges:
            return
        copied_text = ""
        for selection in selected_ranges:
            for row in range(selection.topRow(), selection.bottomRow() + 1):
                row_data = []
                for col in range(selection.leftColumn(), selection.rightColumn() + 1):
                    item = self.order_table.item(row, col)
                    row_data.append(item.text() if item else "")
                copied_text += "\t".join(row_data) + "\n"
        clipboard = QApplication.clipboard()
        clipboard.setText(copied_text)

    def get_filtered_and_sorted_purchase_orders(self):
        filtered_orders = purchase_orders.copy()

        filter_field = self.filter_field_combo.currentText()
        filter_value = self.filter_field_input.text().strip()

        if filter_value:
            pattern = None
            if '*' in filter_value:
                pattern = re.escape(filter_value).replace(r'\*', '.*')
                regex = re.compile(pattern, re.IGNORECASE)
            def match_field(val):
                if pattern:
                    return bool(regex.search(str(val)))
                else:
                    return filter_value.lower() in str(val).lower()

            if filter_field == "按订单号":
                filtered_orders = [o for o in filtered_orders if match_field(o.get('Order Nb',''))]
            elif filter_field == "按Supplier":
                filtered_orders = [o for o in filtered_orders if match_field(o.get('Supplier',''))]
            elif filter_field == "按ITEM Name":
                filtered_orders = [o for o in filtered_orders if match_field(o.get('ITEM Name',''))]
            elif filter_field == "按CATEGORY":
                filtered_orders = [o for o in filtered_orders if match_field(o.get('CATEGORY',''))]
            elif filter_field == "按Date":
                filtered_orders = [o for o in filtered_orders if match_field(o.get('date',''))]
            elif filter_field == "按SKU CLS":
                filtered_orders = [o for o in filtered_orders if match_field(o.get('SKU CLS',''))]

        sort_option = self.sort_combo.currentText()
        def sort_key(order):
            if sort_option == "按更新时间":
                return order.get('date', '') or ''
            elif sort_option == "按订单号":
                return order.get('Order Nb','') or ''
            elif sort_option == "按Product_ID":
                return order.get('Product_ID','') or ''
            elif sort_option == "按SKU CLS":
                return order.get('SKU CLS','') or ''
            elif sort_option == "按Supplier":
                return order.get('Supplier','') or ''
            elif sort_option == "按ITEM Name":
                return order.get('ITEM Name','') or ''
            elif sort_option == "按CATEGORY":
                return order.get('CATEGORY','') or ''
            elif sort_option == "按INVOICE Price":
                try:
                    return float(order.get('INVOICE PRICE',0))
                except:
                    return 0
            elif sort_option == "按WHOLESALE CS":
                try:
                    return float(order.get('WHOLESALE CS',0))
                except:
                    return 0
            elif sort_option == "按EXW":
                try:
                    return float(order.get('EXW EURO',0))
                except:
                    return 0
            return ''

        filtered_orders = sorted(filtered_orders, key=sort_key)
        return filtered_orders

    def on_order_selected(self, selected, deselected):
        try:
            indexes = self.order_table.selectionModel().selectedIndexes()
            if indexes:
                row = indexes[0].row()  # 取第一个选中单元格所在的行
                fs_orders = self.get_filtered_and_sorted_purchase_orders()
                order = fs_orders[row]
                self.current_order = fs_orders[row]
                for field_name, entry in self.entries.items():
                    value = order.get(field_name, "")
                    if isinstance(entry, QComboBox):
                        combo_index = entry.findText(str(value))
                        if combo_index >= 0:
                            entry.setCurrentIndex(combo_index)
                        else:
                            entry.setCurrentText(str(value))
                    else:
                        entry.setText(str(value))
            else:
                for entry in self.entries.values():
                    if isinstance(entry, QComboBox):
                        entry.setCurrentIndex(-1)
                    else:
                        entry.clear()
        except Exception as e:
            print(f"处理订单选择时发生错误：{e}")
            QMessageBox.critical(self, "错误", f"处理订单选择时发生错误：{e}")

    def add_order(self):
        try:
            new_order = {}
            for field_name, entry in self.entries.items():
                if isinstance(entry, QComboBox):
                    value = entry.currentText().strip()
                else:
                    value = entry.text().strip()

                # 数据校验逻辑调整：SKU CLS 应为字符串，不进行整数转换
                # QUANTITY CS 和 BTL PER CS 为整数字段
                # ITEM Name 必填
                # 其他浮点字段进行浮点验证


                if field_name in ["QUANTITY CS", "BTL PER CS", "SKU CLS"]:
                    if not value:
                        QMessageBox.warning(self, "输入错误", f"{field_name} 不能为空！")
                        return
                    try:
                        value = int(value)
                    except ValueError:
                        QMessageBox.warning(self, "输入错误", f"{field_name} 必须是一个有效的整数！")
                        return

                elif field_name == "ITEM Name":
                    if not value:
                        QMessageBox.warning(self, "输入错误", "ITEM Name 不能为空！")
                        return

                elif field_name in ["ALC.", "EXW EURO", "Expected Profit", "Domestic Freight CAD", "International Freight EURO", "EXW Exchange Rate", "International Freight Exchange Rate"]:
                    if not value:
                        QMessageBox.warning(self, "输入错误", f"{field_name} 不能为空！")
                        return
                    try:
                        value = float(value)
                    except ValueError:
                        QMessageBox.warning(self, "输入错误", f"{field_name} 必须是一个有效的数字！")
                        return

                new_order[field_name] = value

            # 添加日期
            new_order['date'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            new_order['PO_GUID'] = str(uuid.uuid4())
            new_order['Order Step'] = '未下单'  # 新订单默认状态
            # 基本检查
            if not new_order['Order Nb']:
                QMessageBox.warning(self, "添加失败", "订单号不能为空！")
                return

            product_id = new_order.get('Product_ID', '')
            if not product_id:
                QMessageBox.warning(self, "添加失败", "产品编号不能为空！")
                return

            if any(order['Order Nb'] == new_order['Order Nb'] for order in purchase_orders):
                QMessageBox.warning(self, "添加失败", "该订单号已存在！")
                return

            # 在添加订单前，先检查 SKU CLS 对应的产品是否存在
            sku_cls = str(new_order.get('SKU CLS', '')).strip()
            product = get_product_by_sku(sku_cls)
            if not product:
                # 产品不存在，尝试添加产品
                if not self.attempt_add_product_to_management(new_order):
                    # 添加产品失败，直接返回
                    return
                else:
                    # 产品添加成功或者已存在了，现在继续添加订单
                    pass

            # 此时产品已存在，可以继续更新库存和保存订单
            quantity_cs = int(new_order.get('QUANTITY CS', 0))
            btl_per_cs = int(new_order.get('BTL PER CS', 0))
            arrival_date = new_order.get('Arrival_Date', '')
            creation_date = new_order.get('date', '')
            item_name = new_order.get('ITEM Name', '')
            sku_cls = str(new_order.get('SKU CLS', '')).strip()
            po_guid = new_order.get('PO_GUID', '')

            update_inventory(
                po_guid,
                product_id,
                new_order['Order Nb'],
                quantity_cs,
                arrival_date,
                creation_date,
                item_name,
                sku_cls,
                btl_per_cs,
                operation_type='add_purchase_order'
            )

            purchase_orders.append(new_order)
            save_purchase_order_to_db(new_order)
            data_manager.data_changed.emit()

            self.update_order_table()
            QMessageBox.information(self, "成功", f"订单 {new_order['Order Nb']} 已添加。")

        except Exception as e:
            print(f"添加订单时发生错误：{e}")
            QMessageBox.critical(self, "错误", f"添加订单时发生错误：{e}")

    def attempt_add_product_to_management(self, order_data):
        """尝试将订单中的产品信息添加到产品列表中（product management）。
           如果添加成功返回 True，失败返回 False。"""
        try:
            # 从订单中获取产品必需信息
            sku_cls = str(order_data.get('SKU CLS', '')).strip()
            item_name = str(order_data.get('ITEM Name', '')).strip()
            category = str(order_data.get('CATEGORY', '')).strip()
            size_str = str(order_data.get('SIZE', '')).strip()
            alc_str = str(order_data.get('ALC.', '')).strip()
            btl_per_cs_str = str(order_data.get('BTL PER CS', '')).strip()
            supplier = str(order_data.get('Supplier', '')).strip()

            # 检查必填字段
            if not sku_cls or not item_name or not category or not size_str or not alc_str or not btl_per_cs_str or not supplier:
                QMessageBox.warning(self, "添加产品失败", "产品信息不完整，无法添加到产品列表中。")
                return False

            # 转换数据类型
            try:
                size = float(size_str)
                alc = float(alc_str)
                btl_per_cs = int(btl_per_cs_str)
            except ValueError:
                QMessageBox.warning(self, "添加产品失败", "Size、ALC、BTL PER CS 必须是有效的数字！")
                return False

            # 检查 SKU_CLS 是否已存在
            if get_product_by_sku(sku_cls) is not None:
                # 已存在，无需重复添加
                return True

            new_product = {
                'SKU_CLS': sku_cls,
                'ITEM_Name': item_name,
                'Category': category,
                'Size': size,
                'ALC': alc,
                'BTL_PER_CS': btl_per_cs,
                'Supplier': supplier,
                'Creation_Date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

            save_product_to_db(new_product)

            # 刷新产品列表
            load_products_from_db()

            # 确认添加成功
            if get_product_by_sku(sku_cls) is not None:
                QMessageBox.information(self, "提示", f"产品 {sku_cls} 已成功添加到产品列表中。")
                return True
            else:
                QMessageBox.warning(self, "添加产品失败", "无法在产品列表中找到新添加的产品，请检查数据。")
                return False

        except Exception as e:
            print(f"尝试添加产品到产品管理时发生错误：{e}")
            QMessageBox.critical(self, "错误", f"添加产品到产品列表时发生错误：{e}")
            return False

    def update_order(self):
        try:
            # 1. 判断是否有选中的订单（缓存的订单记录）
            if not hasattr(self, "current_order") or self.current_order is None:
                QMessageBox.warning(self, "更新失败", "请先选择要更新的订单！")
                return

            # 2. 使用缓存的订单记录作为原始订单
            existing_order = self.current_order
            old_order_nb = existing_order.get("Order Nb", "")
            po_guid = existing_order.get("PO_GUID", "")
            if not po_guid:
                QMessageBox.warning(self, "更新失败", "该订单缺少PO_GUID，无法进行更新！")
                return

            # 3. 从各输入框中读取用户修改后的数据，构造更新后的订单数据
            updated_order = {}
            for field_name, entry in self.entries.items():
                if isinstance(entry, QComboBox):
                    value = entry.currentText().strip()
                else:
                    value = entry.text().strip()

                
                # 对于需要整数转换的字段
                if field_name in ["QUANTITY CS", "BTL PER CS", "SKU CLS"]:
                    if not value:
                        QMessageBox.warning(self, "输入错误", f"{field_name} 不能为空！")
                        return
                    try:
                        value = int(value)
                    except ValueError:
                        QMessageBox.warning(self, "输入错误", f"{field_name} 必须是一个有效的整数！")
                        return
                # 对于需要浮点转换的字段
                elif field_name in ["ALC.", "EXW EURO", "Expected Profit", "Domestic Freight CAD",
                                    "International Freight EURO", "EXW Exchange Rate", "International Freight Exchange Rate"]:
                    if not value:
                        QMessageBox.warning(self, "输入错误", f"{field_name} 不能为空！")
                        return
                    try:
                        value = float(value)
                    except ValueError:
                        QMessageBox.warning(self, "输入错误", f"{field_name} 必须是一个有效的数字！")
                        return
                updated_order["PO_GUID"] = existing_order.get("PO_GUID", "")
                updated_order[field_name] = value

            # 4. 更新日期字段
            updated_order['date'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 5. 在此计算数量差异，调用库存更新函数
            product_id = updated_order.get('Product_ID', '')
            arrival_date = updated_order.get('Arrival_Date', '')
            creation_date = updated_order.get('date', '')
            item_name = updated_order.get('ITEM Name', '')
            sku_cls = updated_order.get('SKU CLS', '')
            btl_per_cs = updated_order.get('BTL PER CS', 0)


            old_quantity_cs = int(existing_order.get('QUANTITY CS', 0))
            new_quantity_cs = int(updated_order.get('QUANTITY CS', 0))
            delta_quantity_cs = new_quantity_cs - old_quantity_cs

            update_inventory(
                po_guid,
                product_id,
                old_order_nb,  # 此处用原始订单号更新库存
                delta_quantity_cs,
                arrival_date,
                creation_date,
                item_name,
                sku_cls,
                btl_per_cs,
                operation_type='update_purchase_order'
            )

            # 6. 获取用户修改后的新订单号（如果用户修改了订单号，则需要级联更新）
            new_order_data = {}

            if updated_order.get("Order Nb") != existing_order.get("Order Nb"):
                new_order_data["Order Nb"] = updated_order["Order Nb"]
            if updated_order.get("Product_ID") != existing_order.get("Product_ID"):
                new_order_data["Product_ID"] = updated_order["Product_ID"]
            if updated_order.get("ITEM Name") != existing_order.get("ITEM Name"):
                new_order_data["ITEM Name"] = updated_order["ITEM Name"]
            if updated_order.get("BTL PER CS") != existing_order.get("BTL PER CS"):
                new_order_data["BTL PER CS"] = updated_order["BTL PER CS"]

            if not updated_order.get("Order Nb", "").strip():
                QMessageBox.warning(self, "输入错误", "订单号不能为空！")
                return           
            new_order_nb = updated_order.get("Order Nb")
            
            if new_order_data:
                cascade_update_purchase_order_by_guid(
                    po_guid=po_guid,
                    new_order_nb=new_order_nb,
                    new_order_data=new_order_data
                )
            # 8. 保存更新后的订单到内存和数据库
            index = purchase_orders.index(existing_order)
            purchase_orders[index] = updated_order
            save_purchase_order_to_db(updated_order)

            # 9. 如有需要，同步更新库存中的 Arrival_Date（这里传入新的订单号）
            update_inventory_arrival_date(product_id, new_order_nb, arrival_date)

            data_manager.data_changed.emit()
            self.update_order_table()
            QMessageBox.information(self, "成功", f"订单 {old_order_nb} 已更新为 {new_order_nb}。")
        except Exception as e:
            print(f"更新订单时发生错误：{e}")
            QMessageBox.critical(self, "错误", f"更新订单时发生错误：{e}")


    
    def check_and_remove_unused_product(self, sku_cls: str):
        """
        判断 products 表中某个 sku_cls 是否还被任何采购订单使用，
        如果没有使用，就把它删除掉。
        """
        try:
            if not sku_cls:
                return
            product = get_product_by_sku(sku_cls)
            if not product:
                return
            # 在 purchase_orders 搜索是否还有其他订单的 SKU CLS == 这个值
            is_used = any(
                (str(o.get('SKU CLS', '')).strip() == sku_cls)
                for o in purchase_orders
            )
            if not is_used:
                # 说明没有别的采购订单在用它，可以安全删除
                delete_product_from_db(sku_cls)
                # 同步从内存的 products 中删除
                product_to_remove = next((p for p in products if p['SKU_CLS'] == sku_cls), None)
                if product_to_remove:
                    products.remove(product_to_remove)
                data_manager.products_changed.emit()
                QMessageBox.information(None, "自动删除旧产品", f"已自动删除旧 SKU_CLS={sku_cls} 对应的产品记录，因为它已不再使用。")
            else:
                QMessageBox.information(None, "旧产品仍然存在", f"请注意！ 旧产品 SKU_CLS={sku_cls} 仍有对应的产品存在。")
        except Exception as e:
            print(f"check_and_remove_unused_product 时发生错误: {e}")
    
    def open_price_calculator(self):
        open_price_calculator(self)

    def find_order(self):
        try:
            search_order_nb = self.entry_search.text().strip()
            if not search_order_nb:
                QMessageBox.warning(self, "查找失败", "请输入要查找的订单号！")
                return
            order = get_purchase_order_by_nb(search_order_nb)
            if order:
                # 在表格中定位并选中该订单
                for row in range(self.order_table.rowCount()):
                    if self.order_table.item(row, 0).text() == search_order_nb:
                        self.order_table.selectRow(row)
                        # 滚动到该行
                        self.order_table.scrollToItem(self.order_table.item(row, 0))
                        break
            else:
                QMessageBox.information(self, "查找结果", f"未找到订单号为 {search_order_nb} 的订单。")
        except Exception as e:
            print(f"查找订单时发生错误：{e}")
            QMessageBox.critical(self, "错误", f"查找订单时发生错误：{e}")


    def delete_order(self):
        try:
            order_nb = self.entry_delete.text().strip()
            if not order_nb:
                QMessageBox.warning(self, "删除失败", "请输入要删除的订单号！")
                return

            # 根据订单号找到采购订单（注意：确保订单号唯一或通过UI选中订单记录）
            order = get_purchase_order_by_nb(order_nb)
            if not order:
                QMessageBox.warning(self, "删除失败", f"未找到订单号为 {order_nb} 的订单。")
                return

            # 获取该订单的 PO_GUID
            po_guid = order.get("PO_GUID", "")
            if not po_guid:
                QMessageBox.warning(self, "删除失败", "该订单缺少 PO_GUID，无法删除！")
                return

            # 1. 检查销售订单引用：遍历所有销售订单，如果有销售订单引用该 PO_GUID，则不允许删除
            referencing_sales = []
            for s_order in sales_orders:
                # 假设销售订单中的 PO_GUID 字段为逗号分隔字符串
                po_guid_list = [guid.strip() for guid in s_order.get("PO_GUID", "").split(',') if guid.strip()]
                if po_guid in po_guid_list:
                    referencing_sales.append(s_order.get("Sales_ID"))
            if referencing_sales:
                QMessageBox.warning(self, "删除错误", f"该采购订单已被销售订单 {', '.join(referencing_sales)} 引用，无法删除！")
                return

            # 2. 检查库存记录状态：通过 PO_GUID 找到库存记录，比较当前库存与原始库存（假设原始库存保存在采购订单中的 "QUANTITY CS" 字段）
            try:
                import sqlite3
                conn = sqlite3.connect(r'D:\00_Programming\98_Pycharm\00_Workplace\Order Manager\Official_Tool\01_Cursor_Code\01_Working\00_Main Branch\orders.db')
                cursor = conn.cursor()
                cursor.execute('SELECT "Current_Stock_CS" FROM inventory WHERE "PO_GUID" = ?', (po_guid,))
                row = cursor.fetchone()
                conn.close()
            except Exception as e:
                print(f"查询库存时发生错误：{e}")
                QMessageBox.critical(self, "错误", f"查询库存时发生错误：{e}")
                return

            if row:
                current_stock = int(row[0])
                original_qty = int(order.get("QUANTITY CS", 0))
                if current_stock < original_qty:
                    QMessageBox.warning(self, "删除错误", "该采购订单已部分售出，无法删除！")
                    return

            # 3. 如果检查通过，执行级联删除：
            #    删除库存记录
            conn = sqlite3.connect(r'D:\00_Programming\98_Pycharm\00_Workplace\Order Manager\Official_Tool\01_Cursor_Code\01_Working\00_Main Branch\orders.db')
            cursor = conn.cursor()
            cursor.execute('DELETE FROM inventory WHERE "PO_GUID" = ?', (po_guid,))
            print(f"已删除库存中 PO_GUID={po_guid} 的记录。")

            #    删除采购订单记录
            cursor.execute('DELETE FROM purchase_orders WHERE "PO_GUID" = ?', (po_guid,))
            print(f"已删除采购订单中 PO_GUID={po_guid} 的记录。")
            conn.commit()
            conn.close()

            # 4. 从内存中删除该订单记录
            global purchase_orders
            purchase_orders = [po for po in purchase_orders if po.get("PO_GUID") != po_guid]
            data_manager.data_changed.emit()
            data_manager.inventory_changed.emit()

            QMessageBox.information(None, "成功", f"采购订单 (PO_GUID={po_guid}) 已删除。")
        except Exception as e:
            print(f"删除采购订单时发生错误：{e}")
            QMessageBox.critical(None, "删除错误", f"删除采购订单时发生错误：{e}")


    def undo_delete_order(self):
        try:
            if not deleted_orders:
                QMessageBox.information(self, "撤销删除", "没有可撤销的删除操作。")
                return
            # 取出最后一个被删除的订单
            order = deleted_orders.pop()
            # 添加回采购订单列表，并保存到数据库
            purchase_orders.append(order)
            save_purchase_order_to_db(order)
            data_manager.data_changed.emit()

            # 对库存进行"恢复"操作：增加库存。假设恢复库存的操作在 add_purchase_order 的库存更新中完成
            quantity_cs = int(order.get("QUANTITY CS", 0))
            btl_per_cs = int(order.get("BTL PER CS", 0))
            product_id = order.get("Product_ID", "")
            order_nb = order.get("Order Nb", "")
            po_guid = get_po_guid_for_inventory(product_id, order_nb)
            update_inventory(
                po_guid,
                product_id,
                order_nb,
                quantity_cs,
                order.get("Arrival_Date", ""),
                order.get("date", ""),
                order.get("ITEM Name", ""),
                order.get("SKU CLS", ""),
                btl_per_cs,
                operation_type='add_purchase_order'
            )
            self.update_order_table()
            QMessageBox.information(None, "成功", f"订单 {order['Order Nb']} 已恢复。")
        except Exception as e:
            print(f"撤销删除时发生错误：{e}")
            QMessageBox.critical(None, "错误", f"撤销删除时发生错误：{e}")



    def export_orders(self):
        try:
            # 获取当前显示的订单(过滤+排序后的列表)
            filtered_sorted_orders = self.get_filtered_and_sorted_purchase_orders()

            if len(filtered_sorted_orders) == 0:
                QMessageBox.information(self, "导出", "没有可导出的订单数据。")
                return

            import pandas as pd
            df = pd.DataFrame(filtered_sorted_orders)

            # 指定导出字段顺序(与原先代码保持一致)
            field_order = [field_name for label_text, field_name in db_fields if field_name in df.columns]
            df = df[field_order]

            # 删除 "PROFIT PER BT" 字段（如果存在）
            if "PROFIT PER BT" in df.columns:
                df.drop("PROFIT PER BT", axis=1, inplace=True, errors='ignore')

            # 尝试将可转为数值的列转为数值类型
            # 这样在输出Excel时，这些列会是数字格式（而非纯文本）
            for col in df.columns:
                try:
                    # 尝试将列转为数值类型，如果失败将触发异常
                    df[col] = pd.to_numeric(df[col])
                except ValueError:
                    # 转换失败表示该列并非纯数字列，保持原样即可，不进行处理
                    pass

            # 弹出文件保存对话框
            from PyQt6.QtWidgets import QFileDialog
            options = QFileDialog.Option.DontUseNativeDialog
            file_name, _ = QFileDialog.getSaveFileName(self, "保存文件", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

            if file_name:
                if not file_name.endswith('.xlsx'):
                    file_name += '.xlsx'
                df.to_excel(file_name, index=False)
                QMessageBox.information(self, "导出成功", f"采购订单已导出到文件 {file_name}")
        except Exception as e:
            print(f"导出订单时发生错误：{e}")
            QMessageBox.critical(self, "错误", f"导出订单时发生错误：{e}")


    def update_order_table(self):
        """
        更新订单表格显示，根据过滤排序后的订单列表更新各单元格，
        并针对价格计算相关字段设置为可编辑，其它字段设置为只读。
        """
        display_fields = [field for field in db_fields]
        filtered_sorted_orders = self.get_filtered_and_sorted_purchase_orders()

        # 在更新前屏蔽 itemChanged 信号，防止内部修改触发信号
        self.order_table.blockSignals(True)
        self.order_table.setRowCount(0)
        self.order_table.setRowCount(len(filtered_sorted_orders))

        # 定义哪些字段允许手动编辑
        editable_fields = {"INVOICE PRICE", "INVOICE CS", "WHOLESALE BTL", "WHOLESALE CS",
                           "TOTAL Freight", "PROFIT PER BT", "PROFIT PER CS", "PROFIT TOTAL",
                           "QUANTITY BTL", "TOTAL AMOUNT", "TOTAL AMOUNT EURO", "REMARKS"}

        for row, order in enumerate(filtered_sorted_orders):
            for col, (label_text, field_name) in enumerate(display_fields):
                value = order.get(field_name, "")
                item = QTableWidgetItem(str(value))
                # 如果该字段在 editable_fields 中，则允许编辑，否则设置只读
                if field_name in editable_fields:
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                else:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.order_table.setItem(row, col, item)
        self.order_table.blockSignals(False)

    def on_item_changed(self, item):
        """
        当用户修改订单表格中某个单元格后，更新对应订单的数据，并保存到数据库。
        为避免内部更新引发递归调用，使用 editing_in_progress 标志。
        """
        if self.editing_in_progress:
            return

        try:
            self.editing_in_progress = True
            col = item.column()
            # 根据 display_fields 获取字段名
            display_fields = [field for field in db_fields]
            field_name = display_fields[col][1]  # display_fields 元素为 (label_text, field_name)
            # 只对可编辑字段进行处理
            editable_fields = {"INVOICE PRICE", "INVOICE CS", "WHOLESALE BTL", "WHOLESALE CS",
                               "TOTAL Freight", "PROFIT PER BT", "PROFIT PER CS", "PROFIT TOTAL",
                               "QUANTITY BTL", "TOTAL AMOUNT", "TOTAL AMOUNT EURO", "REMARKS"}
            if field_name not in editable_fields:
                return

            # 获取当前修改所在的行
            row = item.row()
            # 获取当前显示的订单列表（经过过滤排序后的列表）
            filtered_sorted_orders = self.get_filtered_and_sorted_purchase_orders()
            if row >= len(filtered_sorted_orders):
                return

            # 修改对应订单中的字段
            order = filtered_sorted_orders[row]
            new_value = item.text().strip()
            order[field_name] = new_value
            # 调用保存到数据库的函数（更新时可直接调用 save_purchase_order_to_db）
            save_purchase_order_to_db(order)
            # 若需要，可以发射数据变化信号刷新界面
            data_manager.data_changed.emit()
        except Exception as e:
            print(f"更新订单单元格时发生错误：{e}")
        finally:
            self.editing_in_progress = False

    def showEvent(self, event):
        super().showEvent(event)
        self.update_order_table()

    def compare_with_registration_file(self):
        try:
            # 获取选定的采购订单
            selected_rows = self.order_table.selectionModel().selectedRows()
            if not selected_rows:
                QMessageBox.warning(self, "提示", "请先选择要比较的采购订单（可多选）。")
                return

            fs_orders = self.get_filtered_and_sorted_purchase_orders()
            selected_orders = []
            for index in selected_rows:
                row = index.row()
                order = fs_orders[row]
                selected_orders.append(order)
            for idx, order in enumerate(selected_orders):
                print(f"第 {idx} 条订单：{order}")


            # 让用户选择注册文件（Excel 格式）
            options = QFileDialog.Option.DontUseNativeDialog
            file_name, _ = QFileDialog.getOpenFileName(
                self, "选择注册文件", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

            if not file_name:
                return

            # 读取注册文件
            registration_df = pd.read_excel(file_name)

            # 将列名统一转换为小写
            registration_df.columns = registration_df.columns.str.lower()

            # 打印列名调试
            #print("读取的注册文件列名：", registration_df.columns.tolist())

            # 确保 'sku cls' 字段存在
            if 'sku cls' not in registration_df.columns:
                QMessageBox.warning(self, "错误", "注册文件中缺少 'SKU CLS' 列。")
                return

            # 将 SKU CLS 转为字符串并统一为小写，便于匹配
            registration_df['sku cls'] = registration_df['sku cls'].astype(str).str.strip().str.lower()

            # 创建一个字典，存储采购订单中的信息，键为 SKU CLS（统一转换为小写）
            po_data = {}
            for order in selected_orders:
                sku_cls = str(order.get('SKU CLS', '')).strip().lower()  # 转换为小写
                po_data[sku_cls] = {
                    'category_po': order.get('CATEGORY', ''),
                    'size_po': order.get('SIZE', ''),
                    'alc_po': order.get('ALC.', ''),
                    'btl per cs_po': order.get('BTL PER CS', ''),
                    'wholesale cs_po': order.get('WHOLESALE CS', ''),
                    'supplier_po': order.get('Supplier', ''),
                    'item name_po': order.get('ITEM Name', ''),
                    'invoice cs_po': order.get('INVOICE CS', '')
                }


                # 打印 po_data
                #print("采购订单数据 po_data：", json.dumps(po_data, indent=2, ensure_ascii=False))


            # 需要比较的字段（统一小写以匹配列名）
            fields_to_compare = ['category', 'size', 'alc', 'btl per cs', 'wholesale cs', 'supplier', 'item name', 'invoice cs' ]

            # 在 DataFrame 中插入采购订单的信息
            for field in fields_to_compare:
                po_field = f"{field}_po"
                # 如果原始注册文件中没有对应的列，跳过
                if field not in registration_df.columns:
                    QMessageBox.warning(self, "错误", f"注册文件中缺少 '{field}' 列。")
                    return
                # 在对应的列右侧插入新列
                col_index = registration_df.columns.get_loc(field) + 1
                registration_df.insert(col_index, po_field, "")

            # 遍历注册文件，填入采购订单的信息并比较
            for idx, row in registration_df.iterrows():
                sku_cls = str(row.get('sku cls', '')).strip().lower()  # 转换为小写
                #print(f"注册文件 SKU CLS: {sku_cls}")
                if sku_cls in po_data:
                    #print(f"匹配到的采购订单信息: {po_data[sku_cls]}")
                    for field in fields_to_compare:
                        po_field = f"{field}_po"
                        po_value = po_data[sku_cls].get(f"{field}_PO", '')  # 从 po_data 获取值
                        #print(f"填入 {field}_PO 的值: {po_value}")
                        registration_df.at[idx, po_field] = po_data[sku_cls].get(f"{field}_po", '')
                else:
                    print(f"未匹配到采购订单信息: {sku_cls}")

            # 保存修改后的 DataFrame 到临时 Excel 文件，以便使用 openpyxl 进行样式处理
            temp_file = 'temp_registration.xlsx'
            registration_df.to_excel(temp_file, index=False)

            # 使用 openpyxl 打开临时文件，处理样式
            wb = load_workbook(temp_file)
            ws = wb.active

            # 定义高亮样式RFA
            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # 比较并高亮不一致的单元格
            for idx, row in registration_df.iterrows():
                excel_row = idx + 2  # DataFrame 的索引从 0 开始，Excel 的行从 1 开始，且有标题行
                sku_cls = str(row.get('sku cls', '')).strip().lower()  # 转换为小写
                if sku_cls in po_data:
                    for field in fields_to_compare:
                        po_field = f"{field}_po"
                        reg_value = str(row.get(field, '')).strip()
                        po_value = str(row.get(po_field, '')).strip()
                        if reg_value != po_value:
                            # 高亮注册文件的原始值
                            col_index_reg = registration_df.columns.get_loc(field) + 1  # DataFrame 列索引从 0 开始
                            cell_reg = ws.cell(row=excel_row, column=col_index_reg)
                            cell_reg.fill = fill
                            # 高亮采购订单的值
                            col_index_po = registration_df.columns.get_loc(po_field) + 1
                            cell_po = ws.cell(row=excel_row, column=col_index_po)
                            cell_po.fill = fill

            # 让用户选择保存文件的位置
            save_file_name, _ = QFileDialog.getSaveFileName(
                self, "保存比较结果", "", "Excel Files (*.xlsx);;All Files (*)", options=options)



            #print("注册文件列名：", registration_df.columns.tolist())
            #print("插入后的列名：", registration_df.columns.tolist())


            if not save_file_name:
                return

            if not save_file_name.endswith('.xlsx'):
                save_file_name += '.xlsx'

            # 保存处理后的工作簿
            wb.save(save_file_name)

            # 删除临时文件
            import os
            os.remove(temp_file)

            QMessageBox.information(self, "成功", f"比较完成，结果已保存到 {save_file_name}")

        except Exception as e:
            print(f"比较时发生错误：{e}")
            QMessageBox.critical(self, "错误", f"比较时发生错误：{e}")
